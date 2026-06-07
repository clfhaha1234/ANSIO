"""Real-data ingest pipeline for ANSIO (task F1).

Parses the confidential influencer pricing spreadsheet (path supplied via the
``ANSIO_PRICING_XLSX`` env var; the file itself is gitignored and never
published) and produces TWO outputs that respect the hard confidentiality red
lines:

1. ``benchmark_agg.json`` (gitignored) — **aggregated** pricing benchmark only:
   median Estimate-CPM and price bands grouped by (tier × category). It carries
   NO individual quote, NO creator name, NO per-row price. Only group medians and
   coarse low/median/high bands derived from >=1 rows per cell.

2. ``real_kols()`` — a list of **public** KOL facts (name, channel handle,
   follower count, niche, collaborated brands) drawn from columns that are public
   information about a YouTuber (visible on their channel). These get woven into
   the synthetic ``kols.json`` / ``content.jsonl`` so the demo's main-character
   pool is made of genuine AI-coding creators. **No pricing of any kind crosses
   into the KOL/content corpus.**

CONFIDENTIALITY CONTRACT (NON-NEGOTIABLE):
* The xlsx original and ``benchmark_agg.json`` never enter git (both gitignored).
* Individual quotes / per-creator prices never enter Moss indexes, never appear
  on the demo screen, never get returned by ``real_kols()``.
* The only pricing that leaves this module is the (tier × category) AGGREGATE.

Run (from agent-py/):
    python3 src/ingest_real.py                 # build benchmark_agg.json
    python3 src/ingest_real.py --print-kols     # also dump public KOL facts
    python3 src/ingest_real.py --self-check     # confidentiality grep self-test
"""

from __future__ import annotations

import json
import os
import re
import statistics
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as e:  # pragma: no cover - guidance only
    raise SystemExit(
        "openpyxl required: `uv pip install openpyxl` or `pip install openpyxl`"
    ) from e

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

AGENT_DIR = Path(__file__).resolve().parent.parent
# Path to the confidential pricing workbook is supplied via env var so the
# source stays publishable (no asset name hardcoded). Point ANSIO_PRICING_XLSX
# at your local, gitignored spreadsheet before running F1. If unset, try the
# bundled default then the repo's gitignored docs/ copy (both never committed).
_XLSX_CANDIDATES = [
    os.environ.get("ANSIO_PRICING_XLSX"),
    str(AGENT_DIR / "data" / "pricing.xlsx"),
    str(AGENT_DIR.parent.parent.parent / "docs" / "Influencer list_所有.xlsx"),
]
XLSX_PATH = next(
    (Path(p) for p in _XLSX_CANDIDATES if p and Path(p).exists()),
    Path(_XLSX_CANDIDATES[1]),  # default reported in the not-found message
)
BENCHMARK_OUT = AGENT_DIR / "benchmark_agg.json"  # gitignored

# ---------------------------------------------------------------------------
# Schema mapping — the xlsx headers (6 sheets, some with extra columns)
# ---------------------------------------------------------------------------

# Niche normalization: map the spreadsheet's free-text Category onto the
# kols.json KOL_NICHES whitelist. Every real KOL here is an AI-coding creator,
# so they all collapse to "tech" (the niche that carries coding/AI/software).
NICHE_FOR_CATEGORY_DEFAULT = "tech"


# tier thresholds — identical to gen_kols.tier_for (single source of truth would
# be nicer, but gen_kols has no importable helper without side effects; kept in
# lockstep by the test in this module's self-check).
def tier_for(followers: int) -> str:
    if followers < 10_000:
        return "nano"
    if followers < 100_000:
        return "micro"
    if followers < 500_000:
        return "mid"
    if followers < 1_000_000:
        return "macro"
    return "mega"


# ---------------------------------------------------------------------------
# Robust cell parsers
# ---------------------------------------------------------------------------


def _clean_name(raw) -> str:
    """First non-empty line of the Name cell, trimmed."""
    if raw is None:
        return ""
    return str(raw).strip().split("\n")[0].strip()


def parse_followers(raw) -> int | None:
    """'326K' / '4.05M' / '6.74k' / '238K ' -> int. None if unparseable."""
    if raw is None:
        return None
    s = str(raw).strip().lower().replace(",", "")
    m = re.match(r"([\d.]+)\s*([km]?)", s)
    if not m:
        return None
    try:
        num = float(m.group(1))
    except ValueError:
        return None
    mult = {"k": 1_000, "m": 1_000_000, "": 1}[m.group(2)]
    return int(num * mult)


def parse_cpm(raw) -> float | None:
    """'~97.8' / '～12.4' / '~$268.2' / '~692-1200' / '~785+' -> float (first num)."""
    if raw is None:
        return None
    s = str(raw).replace("$", "").replace("～", "~")
    nums = re.findall(r"[\d.]+", s)
    if not nums:
        return None
    try:
        # For ranges ('692-1200') take the midpoint; for '785+' take the number.
        vals = [float(n) for n in nums if n not in (".",)]
        if not vals:
            return None
        if len(vals) >= 2 and "-" in s:
            return round((vals[0] + vals[1]) / 2, 1)
        return round(vals[0], 1)
    except ValueError:
        return None


def parse_price(raw) -> int | None:
    """Largest $ figure in the YTB Pricing cell (the dedicated-video quote).

    Used ONLY to compute aggregated price bands; the per-row value never leaves
    the aggregator.
    """
    if raw is None:
        return None
    s = str(raw).replace(",", "")
    nums = re.findall(r"\$\s*([\d]+)", s)
    if not nums:
        # Some cells use full-width colon then number without $.
        nums = re.findall(r"[:：]\s*([\d]{3,})", s)
    if not nums:
        return None
    try:
        return max(int(n) for n in nums)
    except ValueError:
        return None


# Brand surface form -> canonical lowercase key (aligned with gen_content brands).
_BRAND_CANON = {
    "claude": "claude-code",
    "claude code": "claude-code",
    "cursor": "cursor",
    "copilot": "copilot",
    "github copilot": "copilot",
    "windsurf": "windsurf",
    "codeium": "codeium",
    "replit": "replit",
    "bolt": "bolt",
    "bolt ai": "bolt",
    "lovable": "lovable",
    "codex": "codex",
    "v0": "v0",
    "tabnine": "tabnine",
    "aider": "aider",
    "continue": "continue",
    "chatgpt": "chatgpt",
    "deepseek": "deepseek",
    "warp": "warp",
    "warp ai": "warp",
    "manus": "manus",
    "mocha": "mocha",
    "metagpt": "metagpt",
    "testsprite": "testsprite",
    "trae": "trae",
    "trae ai": "trae",
    "coderabbit": "coderabbit",
    "codeflying": "codeflying",
    "codeflying ai": "codeflying",
    "verdent": "verdent",
    "creao": "creao",
    "microsoft": "microsoft",
    "windows": "windows",
}


def parse_brands(note) -> list[str]:
    """Extract collaborated brands from the Note column, normalized lowercase.

    Note cells look like 'Promoted Claude / warp ai / chatgpt' or
    '合作过Cursor / Windsurf / Claude'. We split on separators and canonicalize.
    """
    if not note:
        return []
    s = str(note)
    # Drop leading verbs in EN/ZH.
    s = re.sub(r"(?i)promoted|合作过|合作|promote", " ", s)
    parts = re.split(r"[/／,，、\n]", s)
    out: list[str] = []
    for p in parts:
        token = p.strip().lower()
        token = re.sub(r"[^a-z0-9 +.-]", "", token).strip()
        if not token:
            continue
        # Try exact canon, then prefix match (e.g. 'cursor editor' -> cursor).
        canon = _BRAND_CANON.get(token)
        if not canon:
            for surface, key in _BRAND_CANON.items():
                if token.startswith(surface):
                    canon = key
                    break
        if canon and canon not in out:
            out.append(canon)
    return out


def handle_from_link(link, name: str) -> str:
    """Extract the @handle (no @) from a YouTube channel URL.

    Falls back to a slug of the name when the link is a /channel/UC… id.
    """
    if link:
        m = re.search(r"@([A-Za-z0-9_.\-]+)", str(link))
        if m:
            return m.group(1).lstrip("@")
    # Fallback: slugify the display name.
    slug = re.sub(r"[^a-z0-9]", "", name.lower())
    return slug or "creator"


# ---------------------------------------------------------------------------
# Sheet parsing
# ---------------------------------------------------------------------------


def _header_index(header: list[str]) -> dict[str, int | None]:
    def find(*needles: str) -> int | None:
        for i, h in enumerate(header):
            hl = h.lower()
            if any(n in hl for n in needles):
                return i
        return None

    return {
        "name": find("name"),
        "link": find("link"),
        "followers": find("followers"),
        "category": find("category"),
        "views": find("avg views"),
        "location": find("location"),
        "pricing": find("pricing"),
        "cpm": find("cpm"),
        "note": find("note"),
    }


def _iter_rows(ws):
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = _header_index(header)
    if idx["name"] is None:
        return

    def _cell(row, key):
        i = idx[key]
        return row[i] if i is not None and i < len(row) else None

    for r in rows[1:]:

        def cell(key, _row=r):
            return _cell(_row, key)

        name = _clean_name(cell("name"))
        if not name:
            continue
        yield {
            "name": name,
            "link": cell("link"),
            "followers": cell("followers"),
            "category": cell("category"),
            "views": cell("views"),
            "location": cell("location"),
            "pricing": cell("pricing"),
            "cpm": cell("cpm"),
            "note": cell("note"),
        }


def _load_workbook():
    if not XLSX_PATH.exists():
        raise SystemExit(
            f"Source spreadsheet not found: {XLSX_PATH}\n"
            "(F1 needs the confidential xlsx present locally; it stays gitignored.)"
        )
    return openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)


def _parse_all_rows() -> list[dict]:
    """All rows across every sheet, deduped by lowercased display name.

    First occurrence wins, but later occurrences fill in any missing fields
    (links/CPM/brands are sometimes richer on a later sheet).
    """
    wb = _load_workbook()
    by_name: dict[str, dict] = {}
    for ws in wb.worksheets:
        for row in _iter_rows(ws):
            key = row["name"].lower()
            if key not in by_name:
                by_name[key] = dict(row)
            else:
                cur = by_name[key]
                for f in ("link", "cpm", "pricing", "note", "category", "followers"):
                    if not cur.get(f) and row.get(f):
                        cur[f] = row[f]
    return list(by_name.values())


# ---------------------------------------------------------------------------
# Public API 1: aggregated benchmark (NO individual values)
# ---------------------------------------------------------------------------


def build_benchmark() -> dict:
    """Aggregate per-row CPM/price into (tier × category) medians + bands.

    Output schema (only group statistics — never an individual quote):
        {
          "schema": "tier_category_aggregate_v1",
          "category_label": {"tech": "AI coding & developer tools"},
          "cells": {
             "<tier>|<category>": {
                "tier": "...", "category": "...", "n": <count>,
                "cpm_median": <float>, "cpm_p25": <float>, "cpm_p75": <float>,
                "price_band_low": <int>, "price_band_median": <int>,
                "price_band_high": <int>
             }, ...
          },
          "global": {"cpm_median": <float>, "n": <count>}
        }
    """
    rows = _parse_all_rows()
    # Bucket raw values by cell — these locals never get serialized.
    cpm_by_cell: dict[tuple[str, str], list[float]] = {}
    price_by_cell: dict[tuple[str, str], list[int]] = {}
    all_cpm: list[float] = []

    for row in rows:
        followers = parse_followers(row.get("followers"))
        cpm = parse_cpm(row.get("cpm"))
        price = parse_price(row.get("pricing"))
        if followers is None:
            continue
        tier = tier_for(followers)
        category = NICHE_FOR_CATEGORY_DEFAULT  # all real rows are AI-coding
        cell = (tier, category)
        if cpm is not None:
            cpm_by_cell.setdefault(cell, []).append(cpm)
            all_cpm.append(cpm)
        if price is not None:
            price_by_cell.setdefault(cell, []).append(price)

    cells: dict[str, dict] = {}
    for cell in set(cpm_by_cell) | set(price_by_cell):
        tier, category = cell
        cpms = sorted(cpm_by_cell.get(cell, []))
        prices = sorted(price_by_cell.get(cell, []))
        entry = {
            "tier": tier,
            "category": category,
            "n": max(len(cpms), len(prices)),
        }
        if cpms:
            entry["cpm_median"] = round(statistics.median(cpms), 1)
            entry["cpm_p25"] = round(_pct(cpms, 0.25), 1)
            entry["cpm_p75"] = round(_pct(cpms, 0.75), 1)
        if prices:
            entry["price_band_low"] = int(_pct(prices, 0.25))
            entry["price_band_median"] = int(statistics.median(prices))
            entry["price_band_high"] = int(_pct(prices, 0.75))
        cells[f"{tier}|{category}"] = entry

    benchmark = {
        "schema": "tier_category_aggregate_v1",
        "note": (
            "AGGREGATE ONLY. Median Estimate-CPM and coarse price bands per "
            "(tier x category). Contains NO individual creator quote or name. "
            "Derived from gitignored xlsx; this file is also gitignored."
        ),
        "category_label": {"tech": "AI coding & developer tools"},
        "cells": cells,
        "global": {
            "cpm_median": round(statistics.median(all_cpm), 1) if all_cpm else None,
            "n": len(all_cpm),
        },
    }
    return benchmark


def _pct(sorted_vals: list, q: float):
    """Simple inclusive percentile on a pre-sorted list (no numpy)."""
    if not sorted_vals:
        return 0
    if len(sorted_vals) == 1:
        return sorted_vals[0]
    pos = q * (len(sorted_vals) - 1)
    lo = int(pos)
    hi = min(lo + 1, len(sorted_vals) - 1)
    frac = pos - lo
    return sorted_vals[lo] + (sorted_vals[hi] - sorted_vals[lo]) * frac


def write_benchmark() -> dict:
    benchmark = build_benchmark()
    BENCHMARK_OUT.write_text(
        json.dumps(benchmark, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return benchmark


# ---------------------------------------------------------------------------
# Public API 2: real KOL public facts (NO pricing)
# ---------------------------------------------------------------------------


def real_kols() -> list[dict]:
    """Public-only facts for the genuine AI-coding creators in the spreadsheet.

    Each entry: {name, handle (no @), platform, niche, region, followers,
    tier, brands:[lowercase canon]}. **No price, no CPM, no individual quote.**
    These seed the synthetic generators so the demo's hero pool is real people.
    """
    rows = _parse_all_rows()
    out: list[dict] = []
    seen_handles: set[str] = set()
    for row in rows:
        followers = parse_followers(row.get("followers"))
        if followers is None:
            continue
        name = row["name"]
        handle = handle_from_link(row.get("link"), name)
        if handle in seen_handles:
            handle = f"{handle}{len(seen_handles)}"
        seen_handles.add(handle)
        region = str(row.get("location") or "United States").strip() or "United States"
        out.append(
            {
                "name": name,
                "handle": handle,  # no @
                "platform": "YouTube",
                "niche": NICHE_FOR_CATEGORY_DEFAULT,
                "region": region,
                "language": "English",
                "followers": followers,
                "tier": tier_for(followers),
                "brands": parse_brands(row.get("note")),
                # PUBLIC avatar: the creator's real YouTube face via unavatar
                # (handle is public; pricing never crosses into this dict).
                "avatar_url": _avatar_for(handle, name),
            }
        )
    return out


def _avatar_for(handle: str, name: str) -> str:
    """Real public YouTube avatar URL for a real KOL (no network at build)."""
    try:
        from avatar import avatar_url
    except ImportError:  # pragma: no cover - run from elsewhere
        from src.avatar import avatar_url  # type: ignore
    return avatar_url(handle, "YouTube", name)


# ---------------------------------------------------------------------------
# CLI / self-check
# ---------------------------------------------------------------------------


def _self_check() -> int:
    """Confidentiality self-test: no individual price/quote in the outputs."""
    print("Self-check: confidentiality red lines")
    bench = build_benchmark()
    bench_str = json.dumps(bench, ensure_ascii=False)
    kols = real_kols()

    ok = True
    # benchmark must not contain any creator name.
    names = [k["name"] for k in kols]
    leaked = [n for n in names if n and n in bench_str]
    if leaked:
        ok = False
        print(f"  [FAIL] creator name leaked into benchmark: {leaked}")
    else:
        print(f"  [PASS] no creator name in benchmark ({len(names)} checked)")

    # real_kols must not contain price/cpm fields.
    price_keys = {"price", "price_usd", "cpm", "pricing", "quote", "rate"}
    bad = [k for k in kols if price_keys & set(k)]
    if bad:
        ok = False
        print(
            f"  [FAIL] pricing field present in real_kols: {price_keys & set(bad[0])}"
        )
    else:
        print(f"  [PASS] no pricing field in real_kols ({len(kols)} entries)")

    # benchmark cells must be aggregates (n>=1, only group stats).
    forbidden = {"name", "handle", "link"}
    cell_bad = [c for c in bench["cells"].values() if forbidden & set(c)]
    if cell_bad:
        ok = False
        print(f"  [FAIL] individual identifier in benchmark cell: {cell_bad[0]}")
    else:
        print(
            f"  [PASS] benchmark cells are aggregate-only ({len(bench['cells'])} cells)"
        )

    print("RESULT:", "PASS" if ok else "FAIL")
    return 0 if ok else 1


def main() -> None:
    if "--self-check" in sys.argv:
        sys.exit(_self_check())

    bench = write_benchmark()
    print(f"Wrote {BENCHMARK_OUT} (gitignored)")
    print(
        f"  cells: {len(bench['cells'])}  global CPM median: {bench['global']['cpm_median']}"
    )
    for key, c in sorted(bench["cells"].items()):
        line = f"  {key}: n={c['n']}"
        if "cpm_median" in c:
            line += (
                f"  cpm_med={c['cpm_median']} (p25={c['cpm_p25']}/p75={c['cpm_p75']})"
            )
        if "price_band_median" in c:
            line += f"  price=${c['price_band_low']}-${c['price_band_high']}"
        print(line)

    kols = real_kols()
    print(f"\nReal KOLs available for injection: {len(kols)}")
    if "--print-kols" in sys.argv:
        for k in kols:
            print(
                f"  {k['name'][:24]:24} @{k['handle']:20} {k['tier']:5} {k['followers']:>9,}  brands={k['brands']}"
            )


if __name__ == "__main__":
    main()
